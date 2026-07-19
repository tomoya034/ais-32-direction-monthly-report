from __future__ import annotations

import argparse
import calendar
import concurrent.futures
import heapq
import json
import math
import multiprocessing
import os
import queue
import re
import shutil
import statistics
import struct
import subprocess
import sys
import threading
import time
import traceback
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Sequence

import openpyxl
import xlsxwriter


APP_TITLE = "AIS 32方位月報一鍵製作"
APP_VERSION = "0.0.1"
CACHE_VERSION = 2
LEGACY_SPOOL_VERSION = 1
EXCEL_MAX_DATA_ROWS = 1_048_575
LEGACY_SPOOL_MAGIC = b"AISLEG2\0"
LEGACY_RECORD = struct.Struct("<dd")
LEGACY_COUNTS = struct.Struct("<32Q")

DIRECTION_ORDER = [
    "北", "北微東", "北北東", "東北微北", "東北", "東北微東", "東北東", "東微北", "東",
    "東微南", "東南東", "東南微南", "東南", "東南微西", "南南東", "南微東", "南",
    "南微西", "南南西", "西南微南", "西南", "西南微西", "西南西", "西微南", "西",
    "西微北", "西北西", "西北微西", "西北", "西北微北", "北北西", "北微西",
]

LEGACY_DIRECTION_ABBREVIATIONS = [
    "N", "NbE", "NNE", "NEbN", "NE", "NEbE", "ENE", "EbN", "E", "EbS", "ESE",
    "SEbE", "SE", "SEbS", "SSE", "SbE", "S", "SbW", "SSW", "SWbS", "SW", "SWbW",
    "WSW", "WbS", "W", "WbN", "WNW", "NWbW", "NW", "NWbN", "NNW", "NbW",
]

# 影片所示：北至東南東，以及西南西至北微西，共 21 個海向方位。
OPEN_SEA_INDEXES = tuple(range(0, 11)) + tuple(range(22, 32))
OPEN_SEA_DIRECTIONS = tuple(DIRECTION_ORDER[index] for index in OPEN_SEA_INDEXES)
COASTAL_REVIEW_DIRECTIONS = {"西南西", "西微南", "西"}

FILENAME_RE = re.compile(
    r"^D&TMOK\s+KLNG_(?P<date>\d{8})(?:_[^.]*)?\.xlsx$",
    re.IGNORECASE,
)

ProgressCallback = Callable[[dict], None]


class CancelledError(RuntimeError):
    pass


class SourceFileError(RuntimeError):
    pass


@dataclass(frozen=True)
class Candidate:
    distance: float
    bearing: float
    source_row: int
    rank: int | None = None


@dataclass
class DirectionResult:
    direction: str
    candidates: list[Candidate] = field(default_factory=list)
    selected: float | None = None
    selected_bearing: float | None = None
    selected_rank: int | None = None
    cluster_count: int = 0
    status: str = "無資料"
    reason: str = "找不到符合條件的資料"
    over_cap_count: int = 0
    over_cap_max: float | None = None


@dataclass
class DayResult:
    day: date
    source_file: Path | None
    directions: dict[str, DirectionResult]
    rows_scanned: int = 0
    rows_accepted: int = 0
    rows_invalid: int = 0
    rows_wrong_message: int = 0
    rows_not_east: int = 0
    rows_legacy: int = 0
    elapsed_seconds: float = 0.0
    note: str = ""


@dataclass
class AppConfig:
    input_dir: Path
    output_path: Path
    year: int
    month: int
    legacy_output_path: Path | None = None
    max_distance: float = 500.0
    tolerance: float = 0.10
    cluster_size: int = 3
    top_candidates: int = 50
    message_types: tuple[int, ...] = (1, 2, 3, 18, 19)
    workers: int = 1
    overwrite: bool = False
    max_files: int | None = None

    def validate(self) -> None:
        if not self.input_dir.is_dir():
            raise ValueError(f"來源資料夾不存在：{self.input_dir}")
        if not 2000 <= self.year <= 2100:
            raise ValueError("年份必須介於 2000 到 2100。")
        if not 1 <= self.month <= 12:
            raise ValueError("月份必須介於 1 到 12。")
        if not 0 < self.max_distance <= 5000:
            raise ValueError("距離上限必須大於 0。")
        if not 0 < self.tolerance < 1:
            raise ValueError("群聚容許差必須介於 0% 到 100% 之間。")
        if not 2 <= self.cluster_size <= 20:
            raise ValueError("群聚筆數必須介於 2 到 20。")
        if not self.cluster_size <= self.top_candidates <= 1000:
            raise ValueError("候選保留筆數不可小於群聚筆數。")
        if not self.message_types:
            raise ValueError("至少要指定一個 AIS 訊息類型。")
        if not 1 <= self.workers <= 8:
            raise ValueError("平行檔案數必須介於 1 到 8。")
        if self.output_path.suffix.lower() != ".xlsx":
            raise ValueError("輸出檔必須是 .xlsx。")
        if self.legacy_output_path is not None and self.legacy_output_path.suffix.lower() != ".xlsx":
            raise ValueError("原格式相容版輸出檔必須是 .xlsx。")
        if self.legacy_output_path is not None and self.legacy_output_path.resolve() == self.output_path.resolve():
            raise ValueError("新版與原格式版不可使用同一個輸出檔名。")
        if self.output_path.exists() and not self.overwrite:
            raise FileExistsError(f"輸出檔已存在：{self.output_path}")
        if self.legacy_output_path is not None and self.legacy_output_path.exists() and not self.overwrite:
            raise FileExistsError(f"原格式相容版已存在：{self.legacy_output_path}")


def emit(callback: ProgressCallback | None, **payload: object) -> None:
    if callback:
        callback(payload)


def parse_source_date(filename: str) -> date | None:
    match = FILENAME_RE.match(filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group("date"), "%Y%m%d").date()
    except ValueError:
        return None


def discover_source_files(folder: Path) -> tuple[dict[date, Path], list[str]]:
    selected: dict[date, Path] = {}
    warnings: list[str] = []
    for path in sorted(folder.glob("*.xlsx"), key=lambda item: item.name.casefold()):
        parsed = parse_source_date(path.name)
        if parsed is None:
            continue
        previous = selected.get(parsed)
        if previous is None:
            selected[parsed] = path
            continue
        # 同一天多檔時採用最後修改者，同時留下明確警告。
        winner = max((previous, path), key=lambda item: item.stat().st_mtime)
        loser = path if winner == previous else previous
        selected[parsed] = winner
        warnings.append(f"{parsed:%Y-%m-%d} 有重複檔案，採用 {winner.name}，略過 {loser.name}")
    return selected, warnings


def detect_available_months(folder: Path) -> list[tuple[int, int, int]]:
    files, _ = discover_source_files(folder)
    counts: dict[tuple[int, int], int] = {}
    for parsed in files:
        key = (parsed.year, parsed.month)
        counts[key] = counts.get(key, 0) + 1
    return [(year, month, count) for (year, month), count in sorted(counts.items())]


def degree_to_direction_index(value: object) -> tuple[int, float] | None:
    try:
        degree = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(degree):
        return None
    normalized = degree % 360.0
    index = min(int(normalized // 11.25), 31)
    return index, normalized


def _number(value: object) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _message_type(value: object) -> int | None:
    number = _number(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _normalized_header(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.strip().casefold())


REQUIRED_HEADERS = {
    "msg_type": {"msg_type", "msg type"},
    "longitude_desc": {"longitude_desc", "longitude desc"},
    "bearing": {"bearing"},
    "distance": {"distance in nautical miles", "distance_in_nautical_miles"},
}


def locate_headers(header_row: Sequence[object]) -> dict[str, int]:
    normalized = [_normalized_header(value) for value in header_row]
    result: dict[str, int] = {}
    for canonical, aliases in REQUIRED_HEADERS.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                result[canonical] = index
                break
    missing = [key for key in REQUIRED_HEADERS if key not in result]
    if missing:
        friendly = ", ".join(missing)
        raise ValueError(f"缺少必要欄位：{friendly}")
    return result


def select_cluster(
    direction: str,
    candidates: list[Candidate],
    tolerance: float,
    cluster_size: int,
    over_cap_count: int,
    over_cap_max: float | None,
) -> DirectionResult:
    result = DirectionResult(
        direction=direction,
        candidates=candidates,
        over_cap_count=over_cap_count,
        over_cap_max=over_cap_max,
    )
    if not candidates:
        if over_cap_count:
            result.reason = f"只有超過距離上限的資料（{over_cap_count:,} 筆）"
        return result

    for start, candidate in enumerate(candidates):
        floor = candidate.distance * (1.0 - tolerance)
        count = 0
        for follower in candidates[start:]:
            if follower.distance + 1e-12 < floor:
                break
            count += 1
        if count >= cluster_size:
            result.selected = candidate.distance
            result.selected_bearing = candidate.bearing
            result.selected_rank = start + 1
            result.cluster_count = count
            result.status = "自動採用"
            discarded = start
            result.reason = (
                f"第 {start + 1} 名起有 {count} 筆落在 {tolerance:.0%} 範圍內"
                + (f"；前 {discarded} 筆視為孤立值" if discarded else "")
            )
            if direction in COASTAL_REVIEW_DIRECTIONS and candidate.distance > 10:
                result.status = "待複核"
                result.reason += "；此岸向方位超過影片提示的 10 NM 經驗值"
            return result

    fallback = candidates[0]
    result.selected = fallback.distance
    result.selected_bearing = fallback.bearing
    result.selected_rank = 1
    result.cluster_count = 1
    result.status = "待複核"
    result.reason = f"前 {len(candidates)} 筆中找不到至少 {cluster_size} 筆的 {tolerance:.0%} 群聚，暫用最高值"
    return result


def select_cluster_from_sorted_rows(
    direction: str,
    rows: list[tuple[float, int, float]],
    config: AppConfig,
    over_cap_count: int,
    over_cap_max: float | None,
) -> DirectionResult:
    valid_start = over_cap_count
    candidates = [
        Candidate(distance=item[0], source_row=item[1], bearing=item[2], rank=rank)
        for rank, item in enumerate(
            rows[valid_start : valid_start + config.top_candidates],
            start=1,
        )
    ]
    result = DirectionResult(
        direction=direction,
        candidates=candidates,
        over_cap_count=over_cap_count,
        over_cap_max=over_cap_max,
    )
    valid_count = len(rows) - valid_start
    if valid_count <= 0:
        if over_cap_count:
            result.reason = f"只有超過距離上限的資料（{over_cap_count:,} 筆）"
        return result

    last_possible = len(rows) - config.cluster_size
    for start in range(valid_start, last_possible + 1):
        candidate_distance, _source_row, candidate_bearing = rows[start]
        floor = candidate_distance * (1.0 - config.tolerance)
        required_follower = rows[start + config.cluster_size - 1]
        if required_follower[0] + 1e-12 < floor:
            continue
        stop = start + config.cluster_size
        while stop < len(rows) and rows[stop][0] + 1e-12 >= floor:
            stop += 1
        rank = start - valid_start + 1
        result.selected = candidate_distance
        result.selected_bearing = candidate_bearing
        result.selected_rank = rank
        result.cluster_count = stop - start
        result.status = "自動採用"
        result.reason = (
            f"第 {rank} 名起有 {result.cluster_count} 筆落在 {config.tolerance:.0%} 範圍內"
            + (f"；前 {rank - 1} 筆視為孤立值" if rank > 1 else "")
        )
        if direction in COASTAL_REVIEW_DIRECTIONS and candidate_distance > 10:
            result.status = "待複核"
            result.reason += "；此岸向方位超過影片提示的 10 NM 經驗值"
        if rank > config.top_candidates:
            extra_stop = min(stop, start + config.top_candidates)
            candidates.extend(
                Candidate(
                    distance=item[0],
                    source_row=item[1],
                    bearing=item[2],
                    rank=actual_rank,
                )
                for actual_rank, item in enumerate(rows[start:extra_stop], start=rank)
            )
        return result

    fallback_distance, _source_row, fallback_bearing = rows[valid_start]
    result.selected = fallback_distance
    result.selected_bearing = fallback_bearing
    result.selected_rank = 1
    result.cluster_count = 1
    result.status = "待複核"
    result.reason = f"全部 {valid_count:,} 筆中找不到至少 {config.cluster_size} 筆的 {config.tolerance:.0%} 群聚，暫用最高值"
    return result


def empty_day_result(day: date, note: str = "當日來源檔缺漏") -> DayResult:
    directions = {
        name: DirectionResult(direction=name, status="無資料", reason=note)
        for name in DIRECTION_ORDER
    }
    return DayResult(day=day, source_file=None, directions=directions, note=note)


def _cache_directory(config: AppConfig) -> Path:
    return config.output_path.parent / f".{config.output_path.stem}_cache"


def _legacy_spool_directory(config: AppConfig) -> Path:
    output = config.legacy_output_path or config.output_path
    return output.parent / f".{output.stem}_rows"


def _legacy_spool_path(config: AppConfig, parsed_day: date) -> Path:
    return _legacy_spool_directory(config) / f"{parsed_day:%Y%m%d}.bin"


def _cache_signature(source_file: Path, parsed_day: date, config: AppConfig) -> dict[str, object]:
    stats = source_file.stat()
    return {
        "cache_version": CACHE_VERSION,
        "source": str(source_file.resolve()),
        "size": stats.st_size,
        "modified_ns": stats.st_mtime_ns,
        "date": parsed_day.isoformat(),
        "max_distance": config.max_distance,
        "tolerance": config.tolerance,
        "cluster_size": config.cluster_size,
        "top_candidates": config.top_candidates,
        "message_types": list(config.message_types),
        "legacy_spool_version": LEGACY_SPOOL_VERSION if config.legacy_output_path is not None else None,
    }


def _day_result_to_dict(day_result: DayResult) -> dict[str, object]:
    return {
        "day": day_result.day.isoformat(),
        "source_file": str(day_result.source_file) if day_result.source_file else None,
        "rows_scanned": day_result.rows_scanned,
        "rows_accepted": day_result.rows_accepted,
        "rows_invalid": day_result.rows_invalid,
        "rows_wrong_message": day_result.rows_wrong_message,
        "rows_not_east": day_result.rows_not_east,
        "rows_legacy": day_result.rows_legacy,
        "elapsed_seconds": day_result.elapsed_seconds,
        "note": day_result.note,
        "directions": {
            direction: {
                "selected": result.selected,
                "selected_bearing": result.selected_bearing,
                "selected_rank": result.selected_rank,
                "cluster_count": result.cluster_count,
                "status": result.status,
                "reason": result.reason,
                "over_cap_count": result.over_cap_count,
                "over_cap_max": result.over_cap_max,
                "candidates": [
                    {"distance": item.distance, "bearing": item.bearing, "source_row": item.source_row, "rank": item.rank}
                    for item in result.candidates
                ],
            }
            for direction, result in day_result.directions.items()
        },
    }


def _day_result_from_dict(payload: dict[str, object]) -> DayResult:
    directions: dict[str, DirectionResult] = {}
    raw_directions = payload["directions"]
    if not isinstance(raw_directions, dict):
        raise ValueError("快取 directions 格式錯誤")
    for direction, raw in raw_directions.items():
        if not isinstance(raw, dict):
            raise ValueError("快取 direction 格式錯誤")
        raw_candidates = raw.get("candidates", [])
        candidates = [
            Candidate(
                distance=float(item["distance"]),
                bearing=float(item["bearing"]),
                source_row=int(item["source_row"]),
                rank=int(item["rank"]) if item.get("rank") is not None else None,
            )
            for item in raw_candidates
        ]
        directions[str(direction)] = DirectionResult(
            direction=str(direction),
            candidates=candidates,
            selected=raw.get("selected"),
            selected_bearing=raw.get("selected_bearing"),
            selected_rank=raw.get("selected_rank"),
            cluster_count=int(raw.get("cluster_count", 0)),
            status=str(raw.get("status", "無資料")),
            reason=str(raw.get("reason", "")),
            over_cap_count=int(raw.get("over_cap_count", 0)),
            over_cap_max=raw.get("over_cap_max"),
        )
    source_value = payload.get("source_file")
    return DayResult(
        day=date.fromisoformat(str(payload["day"])),
        source_file=Path(str(source_value)) if source_value else None,
        directions=directions,
        rows_scanned=int(payload.get("rows_scanned", 0)),
        rows_accepted=int(payload.get("rows_accepted", 0)),
        rows_invalid=int(payload.get("rows_invalid", 0)),
        rows_wrong_message=int(payload.get("rows_wrong_message", 0)),
        rows_not_east=int(payload.get("rows_not_east", 0)),
        rows_legacy=int(payload.get("rows_legacy", payload.get("rows_accepted", 0))),
        elapsed_seconds=float(payload.get("elapsed_seconds", 0.0)),
        note=str(payload.get("note", "")),
    )


def load_cached_day(source_file: Path, parsed_day: date, config: AppConfig) -> DayResult | None:
    cache_path = _cache_directory(config) / f"{parsed_day:%Y%m%d}.json"
    if not cache_path.is_file():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
        if payload.get("signature") != _cache_signature(source_file, parsed_day, config):
            return None
        result = _day_result_from_dict(payload["result"])
        if config.legacy_output_path is not None:
            spool_path = _legacy_spool_path(config, parsed_day)
            if not spool_path.is_file() or sum(legacy_spool_counts(spool_path)) != result.rows_legacy:
                return None
        return result
    except (OSError, ValueError, TypeError, KeyError, json.JSONDecodeError):
        return None


def save_cached_day(source_file: Path, parsed_day: date, config: AppConfig, day_result: DayResult) -> None:
    cache_dir = _cache_directory(config)
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{parsed_day:%Y%m%d}.json"
    temp_path = cache_path.with_suffix(".tmp")
    payload = {
        "signature": _cache_signature(source_file, parsed_day, config),
        "result": _day_result_to_dict(day_result),
    }
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    os.replace(temp_path, cache_path)


def write_legacy_spool(
    config: AppConfig,
    parsed_day: date,
    rows: list[list[tuple[float, int, float]]],
    *,
    presorted: bool = False,
) -> Path:
    spool_dir = _legacy_spool_directory(config)
    spool_dir.mkdir(parents=True, exist_ok=True)
    target = _legacy_spool_path(config, parsed_day)
    temporary = target.with_suffix(".tmp")
    counts = [len(bucket) for bucket in rows]
    with temporary.open("wb") as handle:
        handle.write(LEGACY_SPOOL_MAGIC)
        handle.write(LEGACY_COUNTS.pack(*counts))
        for bucket in rows:
            if not presorted:
                bucket.sort(key=lambda item: (-item[0], item[1]))
            for distance, _source_row, bearing in bucket:
                handle.write(LEGACY_RECORD.pack(distance, bearing))
            bucket.clear()
    os.replace(temporary, target)
    return target


def legacy_spool_counts(path: Path) -> list[int]:
    with path.open("rb") as handle:
        if handle.read(len(LEGACY_SPOOL_MAGIC)) != LEGACY_SPOOL_MAGIC:
            raise ValueError(f"舊格式暫存檔格式錯誤：{path.name}")
        counts_raw = handle.read(LEGACY_COUNTS.size)
        if len(counts_raw) != LEGACY_COUNTS.size:
            raise ValueError(f"舊格式暫存檔不完整：{path.name}")
        counts = list(LEGACY_COUNTS.unpack(counts_raw))
    expected_size = len(LEGACY_SPOOL_MAGIC) + LEGACY_COUNTS.size + sum(counts) * LEGACY_RECORD.size
    if path.stat().st_size != expected_size:
        raise ValueError(f"舊格式暫存檔大小不符：{path.name}")
    return counts


def read_legacy_spool(path: Path) -> tuple[list[int], Iterable[tuple[int, float, float]]]:
    handle = path.open("rb")
    try:
        counts = legacy_spool_counts(path)
        handle.seek(len(LEGACY_SPOOL_MAGIC) + LEGACY_COUNTS.size)

        def records() -> Iterable[tuple[int, float, float]]:
            try:
                for direction_index, count in enumerate(counts):
                    for _ in range(count):
                        raw = handle.read(LEGACY_RECORD.size)
                        if len(raw) != LEGACY_RECORD.size:
                            raise ValueError(f"舊格式暫存檔資料中斷：{path.name}")
                        distance, bearing = LEGACY_RECORD.unpack(raw)
                        yield direction_index, distance, bearing
                if handle.read(1):
                    raise ValueError(f"舊格式暫存檔尾端有非預期資料：{path.name}")
            finally:
                handle.close()

        return counts, records()
    except Exception:
        handle.close()
        raise


def _process_file_worker(arguments: tuple[Path, date, AppConfig]) -> tuple[date, DayResult]:
    source_file, parsed_day, config = arguments
    return parsed_day, process_source_file(source_file, parsed_day, config)


def preflight_output_space(config: AppConfig, source_files: Iterable[Path]) -> tuple[int, int | None]:
    targets = [config.output_path]
    if config.legacy_output_path is not None:
        targets.append(config.legacy_output_path)
    for target in targets:
        target.parent.mkdir(parents=True, exist_ok=True)

    source_size = sum(path.stat().st_size for path in source_files)
    if config.legacy_output_path is None:
        required = max(256 * 1024**2, int(source_size * 0.08))
    else:
        required = max(1536 * 1024**2, int(source_size * 0.55))
    try:
        free = shutil.disk_usage(config.output_path.parent).free
    except OSError:
        free = None
    if free is not None and free > 0 and free < required:
        raise OSError(
            f"輸出磁碟空間可能不足。估計至少需要 {required / 1024**3:.1f} GB，"
            f"目前可用 {free / 1024**3:.1f} GB。請更換輸出位置或清出空間。"
        )
    return required, free


def process_source_file(
    source_file: Path,
    parsed_day: date,
    config: AppConfig,
    callback: ProgressCallback | None = None,
    cancel_event: threading.Event | None = None,
) -> DayResult:
    started = time.perf_counter()
    heaps: list[list[tuple[float, int, float]]] = [[] for _ in DIRECTION_ORDER]
    legacy_rows: list[list[tuple[float, int, float]]] | None = (
        [[] for _ in DIRECTION_ORDER] if config.legacy_output_path is not None else None
    )
    over_cap_counts = [0] * len(DIRECTION_ORDER)
    over_cap_maxes: list[float | None] = [None] * len(DIRECTION_ORDER)
    result = DayResult(
        day=parsed_day,
        source_file=source_file,
        directions={},
    )

    try:
        workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as error:
        raise SourceFileError(
            f"無法開啟來源檔「{source_file.name}」。檔案可能損壞、尚未同步完成，或不是有效的 .xlsx。原始錯誤：{error}"
        ) from error
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise SourceFileError(f"來源檔「{source_file.name}」第一張工作表沒有標題列。")
        try:
            indexes = locate_headers(header)
        except ValueError as error:
            raise SourceFileError(
                f"來源檔「{source_file.name}」{error}。需要 msg_type、LONGITUDE_DESC、bearing、distance in nautical miles。"
            ) from error
        max_index = max(indexes.values())
        estimated_rows = max((worksheet.max_row or 1) - 1, 1)

        for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            result.rows_scanned += 1
            if cancel_event and excel_row % 10000 == 0 and cancel_event.is_set():
                raise CancelledError("使用者已取消")
            if len(row) <= max_index:
                result.rows_invalid += 1
                continue

            msg_type = _message_type(row[indexes["msg_type"]])
            if msg_type not in config.message_types:
                result.rows_wrong_message += 1
                continue
            longitude_desc = row[indexes["longitude_desc"]]
            if str(longitude_desc).strip().casefold() != "east":
                result.rows_not_east += 1
                continue

            direction_info = degree_to_direction_index(row[indexes["bearing"]])
            distance = _number(row[indexes["distance"]])
            if direction_info is None or distance is None or distance < 0:
                result.rows_invalid += 1
                continue

            direction_index, normalized_bearing = direction_info
            if legacy_rows is not None:
                result.rows_legacy += 1
                if result.rows_legacy > EXCEL_MAX_DATA_ROWS:
                    raise SourceFileError(
                        f"來源檔「{source_file.name}」符合舊格式的資料超過 Excel 單張工作表上限 {EXCEL_MAX_DATA_ROWS:,} 列。"
                    )
                legacy_rows[direction_index].append((distance, excel_row, normalized_bearing))
            if distance > config.max_distance:
                over_cap_counts[direction_index] += 1
                prior_max = over_cap_maxes[direction_index]
                if prior_max is None or distance > prior_max:
                    over_cap_maxes[direction_index] = distance
                continue

            result.rows_accepted += 1
            heap = heaps[direction_index]
            item = (distance, excel_row, normalized_bearing)
            if len(heap) < config.top_candidates:
                heapq.heappush(heap, item)
            elif item[0] > heap[0][0]:
                heapq.heapreplace(heap, item)

            if excel_row % 100000 == 0:
                emit(
                    callback,
                    kind="rows",
                    file=source_file.name,
                    rows=result.rows_scanned,
                    estimated_rows=estimated_rows,
                )
    finally:
        workbook.close()

    if legacy_rows is not None:
        for index, direction in enumerate(DIRECTION_ORDER):
            legacy_rows[index].sort(key=lambda item: (-item[0], item[1]))
            result.directions[direction] = select_cluster_from_sorted_rows(
                direction=direction,
                rows=legacy_rows[index],
                config=config,
                over_cap_count=over_cap_counts[index],
                over_cap_max=over_cap_maxes[index],
            )
        write_legacy_spool(config, parsed_day, legacy_rows, presorted=True)
    else:
        for index, direction in enumerate(DIRECTION_ORDER):
            ordered = [
                Candidate(distance=item[0], source_row=item[1], bearing=item[2], rank=rank)
                for rank, item in enumerate(
                    sorted(heaps[index], key=lambda value: (-value[0], value[1])),
                    start=1,
                )
            ]
            result.directions[direction] = select_cluster(
                direction=direction,
                candidates=ordered,
                tolerance=config.tolerance,
                cluster_size=config.cluster_size,
                over_cap_count=over_cap_counts[index],
                over_cap_max=over_cap_maxes[index],
            )

    result.elapsed_seconds = time.perf_counter() - started
    return result


def process_month(
    config: AppConfig,
    callback: ProgressCallback | None = None,
    cancel_event: threading.Event | None = None,
) -> tuple[list[DayResult], list[str]]:
    config.validate()
    discovered, warnings = discover_source_files(config.input_dir)
    month_files = {
        parsed: path
        for parsed, path in discovered.items()
        if parsed.year == config.year and parsed.month == config.month
    }
    if not month_files:
        raise ValueError(f"來源資料夾內找不到 {config.year} 年 {config.month} 月的 D&TMOK KLNG 檔案。")

    required, free = preflight_output_space(config, month_files.values())
    emit(callback, kind="preflight", required_bytes=required, free_bytes=free)

    ordered_files = sorted(month_files.items())
    if config.max_files is not None:
        ordered_files = ordered_files[: config.max_files]
        allowed_days = {parsed for parsed, _ in ordered_files}
    else:
        allowed_days = set(month_files)

    processed: dict[date, DayResult] = {}
    total_files = len(ordered_files)
    pending: list[tuple[date, Path]] = []
    completed_count = 0
    for parsed, source_file in ordered_files:
        cached = load_cached_day(source_file, parsed, config)
        if cached is not None:
            processed[parsed] = cached
            completed_count += 1
            emit(
                callback,
                kind="cache_hit",
                position=completed_count,
                total=total_files,
                file=source_file.name,
            )
        else:
            pending.append((parsed, source_file))

    if config.workers == 1 or len(pending) <= 1:
        for parsed, source_file in pending:
            if cancel_event and cancel_event.is_set():
                raise CancelledError("使用者已取消")
            emit(
                callback,
                kind="file_start",
                position=completed_count + 1,
                total=total_files,
                file=source_file.name,
            )
            result = process_source_file(
                source_file,
                parsed,
                config,
                callback=callback,
                cancel_event=cancel_event,
            )
            processed[parsed] = result
            save_cached_day(source_file, parsed, config, result)
            completed_count += 1
            emit(
                callback,
                kind="file_done",
                position=completed_count,
                total=total_files,
                file=source_file.name,
                seconds=result.elapsed_seconds,
            )
    elif pending:
        emit(callback, kind="parallel_start", workers=config.workers, total=len(pending))
        with concurrent.futures.ProcessPoolExecutor(max_workers=config.workers) as executor:
            futures = {
                executor.submit(_process_file_worker, (source_file, parsed, config)): (parsed, source_file)
                for parsed, source_file in pending
            }
            try:
                for future in concurrent.futures.as_completed(futures):
                    if cancel_event and cancel_event.is_set():
                        for other in futures:
                            other.cancel()
                        raise CancelledError("使用者已取消")
                    expected_day, source_file = futures[future]
                    parsed, result = future.result()
                    if parsed != expected_day:
                        raise RuntimeError(f"平行處理日期不一致：{parsed} / {expected_day}")
                    processed[parsed] = result
                    save_cached_day(source_file, parsed, config, result)
                    completed_count += 1
                    emit(
                        callback,
                        kind="file_done",
                        position=completed_count,
                        total=total_files,
                        file=source_file.name,
                        seconds=result.elapsed_seconds,
                    )
            except Exception:
                for future in futures:
                    future.cancel()
                raise

    days_in_month = calendar.monthrange(config.year, config.month)[1]
    output: list[DayResult] = []
    for day_number in range(1, days_in_month + 1):
        parsed = date(config.year, config.month, day_number)
        if parsed in processed:
            output.append(processed[parsed])
        elif config.max_files is not None and parsed in month_files and parsed not in allowed_days:
            output.append(empty_day_result(parsed, "測試模式未處理此檔"))
        else:
            output.append(empty_day_result(parsed))
            warnings.append(f"缺少 {parsed:%Y-%m-%d} 的來源檔")
    return output, warnings


def _excel_col(column_zero_based: int) -> str:
    value = column_zero_based + 1
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", name)[:31]


def _stats(values: Iterable[float | None]) -> tuple[float | None, float | None, float | None, float | None]:
    valid = [float(value) for value in values if value is not None]
    if not valid:
        return None, None, None, None
    stdev = statistics.stdev(valid) if len(valid) >= 2 else None
    return max(valid), min(valid), statistics.mean(valid), stdev


def write_monthly_workbook(
    config: AppConfig,
    day_results: list[DayResult],
    warnings: list[str],
    callback: ProgressCallback | None = None,
    cancel_event: threading.Event | None = None,
) -> Path:
    config.output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = config.output_path.with_name(config.output_path.stem + ".building.xlsx")
    if temp_path.exists():
        temp_path.unlink()

    emit(callback, kind="write_start", file=config.output_path.name)
    workbook = xlsxwriter.Workbook(
        temp_path,
        {
            "nan_inf_to_errors": True,
            "default_date_format": "yyyy-mm-dd",
        },
    )
    workbook.set_properties(
        {
            "title": f"KLNG {config.year}年{config.month}月 AIS 32方位數值",
            "subject": "AIS 通訊距離自動整理與複核",
            "author": "AIS 32方位月報一鍵製作",
            "comments": "由原始 D&TMOK KLNG 每日檔自動產生；原始檔未修改。",
        }
    )
    workbook.set_calc_mode("auto")

    formats = {
        "title": workbook.add_format(
            {"font_name": "Microsoft JhengHei", "font_size": 18, "bold": True, "font_color": "#FFFFFF", "bg_color": "#17324D", "align": "left", "valign": "vcenter"}
        ),
        "section": workbook.add_format(
            {"font_name": "Microsoft JhengHei", "font_size": 12, "bold": True, "font_color": "#17324D", "bg_color": "#DCEAF5", "bottom": 1, "bottom_color": "#8FAFC6"}
        ),
        "header": workbook.add_format(
            {"font_name": "Microsoft JhengHei", "bold": True, "font_color": "#FFFFFF", "bg_color": "#2F6B7C", "align": "center", "valign": "vcenter", "text_wrap": True, "bottom": 1, "bottom_color": "#17324D"}
        ),
        "subheader": workbook.add_format(
            {"font_name": "Microsoft JhengHei", "bold": True, "font_color": "#17324D", "bg_color": "#EAF3F7", "align": "center", "valign": "vcenter", "text_wrap": True}
        ),
        "body": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 10, "valign": "top"}),
        "body_wrap": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 10, "valign": "top", "text_wrap": True}),
        "number": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 10, "num_format": "0.000", "align": "right"}),
        "integer": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 10, "num_format": "#,##0", "align": "right"}),
        "date": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 10, "num_format": "yyyy-mm-dd", "align": "center"}),
        "auto": workbook.add_format({"font_name": "Microsoft JhengHei", "bg_color": "#E6F4EA", "font_color": "#1F6D3A", "num_format": "0.000", "align": "center"}),
        "review": workbook.add_format({"font_name": "Microsoft JhengHei", "bg_color": "#FFF2CC", "font_color": "#8A5A00", "num_format": "0.000", "align": "center"}),
        "missing": workbook.add_format({"font_name": "Microsoft JhengHei", "bg_color": "#FCE8E6", "font_color": "#A61B1B", "align": "center"}),
        "excluded": workbook.add_format({"font_name": "Microsoft JhengHei", "bg_color": "#E7E9EC", "font_color": "#6B7280", "align": "center"}),
        "input": workbook.add_format({"font_name": "Microsoft JhengHei", "bg_color": "#FFF9E6", "font_color": "#7A4B00", "num_format": "0.000", "align": "center", "border": 1, "border_color": "#E0C36C"}),
        "final": workbook.add_format({"font_name": "Microsoft JhengHei", "bold": True, "bg_color": "#DCEAF5", "font_color": "#17324D", "num_format": "0.000", "align": "center"}),
        "note": workbook.add_format({"font_name": "Microsoft JhengHei", "font_size": 9, "font_color": "#5B6573", "italic": True, "text_wrap": True}),
        "stat_label": workbook.add_format({"font_name": "Microsoft JhengHei", "bold": True, "font_color": "#FFFFFF", "bg_color": "#17324D", "align": "center"}),
        "stat": workbook.add_format({"font_name": "Microsoft JhengHei", "bold": True, "num_format": "0.000", "bg_color": "#EDF4F7", "align": "center"}),
        "link": workbook.add_format({"font_name": "Microsoft JhengHei", "font_color": "#0563C1", "underline": True, "align": "center"}),
    }

    guide = workbook.add_worksheet("操作說明")
    guide.hide_gridlines(2)
    guide.set_column("A:A", 22)
    guide.set_column("B:B", 92)
    guide.set_row(0, 34)
    guide.merge_range("A1:B1", APP_TITLE, formats["title"])
    guide.write("A3", "本次設定", formats["section"])
    settings = [
        ("製作月份", f"{config.year} 年 {config.month} 月"),
        ("來源資料夾", str(config.input_dir)),
        ("輸出檔", str(config.output_path)),
        ("訊息類型", ", ".join(map(str, config.message_types))),
        ("經度描述", "LONGITUDE_DESC = East"),
        ("距離上限", f"{config.max_distance:g} NM；超過者不納入自動值並保留計數"),
        ("群聚規則", f"由高至低尋找至少 {config.cluster_size} 筆、彼此位於最高值減 {config.tolerance:.0%} 範圍內的第一群"),
        ("候選保留", f"每個方向保留前 {config.top_candidates} 筆合格資料，供追查與人工覆核"),
        ("處理方位", "北至東南東、西南西至北微西，共 21 方位；朝向臺灣的中間 11 方位不處理"),
        ("原始資料", "所有來源檔只讀取、不修改；清理與覆核結果另存於本月報"),
    ]
    for row, (label, value) in enumerate(settings, start=3):
        guide.write(row, 0, label, formats["subheader"])
        guide.write(row, 1, value, formats["body_wrap"])
        guide.set_row(row, 30 if len(value) > 70 else 22)

    next_row = 3 + len(settings) + 2
    guide.write(next_row, 0, "使用方式", formats["section"])
    instructions = [
        "先看「總表」與兩張圖，這些值會引用各日工作表的「最終值」。",
        "再看「待複核」；黃色項目代表群聚不足，或西南西／西微南／西超過 10 NM。",
        "如要修正，前往該日工作表，在方向欄的「人工覆核值」輸入距離；「最終值」與總表會自動改用人工值。",
        "每日工作表 A:F 保存自動判斷所用的候選資料與原始列號，可回查來源檔。",
        "自動規則是把影片中的人工判斷具體化；它不能取代研究上的最終判斷，因此所有例外均明確標記。",
    ]
    for offset, instruction in enumerate(instructions, start=1):
        guide.write(next_row + offset, 0, f"{offset}.", formats["subheader"])
        guide.write(next_row + offset, 1, instruction, formats["body_wrap"])
        guide.set_row(next_row + offset, 32)

    warning_row = next_row + len(instructions) + 2
    guide.write(warning_row, 0, "檔案警告", formats["section"])
    if warnings:
        for offset, warning in enumerate(warnings, start=1):
            guide.write(warning_row + offset, 0, offset, formats["integer"])
            guide.write(warning_row + offset, 1, warning, formats["body_wrap"])
    else:
        guide.write(warning_row + 1, 1, "無", formats["body"])
    guide.freeze_panes(2, 0)
    guide.set_landscape()
    guide.fit_to_pages(1, 0)

    daily_sheet_names: dict[date, str] = {}
    for day_index, day_result in enumerate(day_results, start=1):
        if cancel_event and cancel_event.is_set():
            workbook.close()
            raise CancelledError("使用者已取消")
        sheet_name = _safe_sheet_name(f"{config.month}月{day_result.day.day}日")
        daily_sheet_names[day_result.day] = sheet_name
        sheet = workbook.add_worksheet(sheet_name)
        sheet.hide_gridlines(2)
        sheet.freeze_panes(1, 0)
        sheet.set_column("A:A", 11)
        sheet.set_column("B:B", 11)
        sheet.set_column("C:C", 11)
        sheet.set_column("D:D", 12)
        sheet.set_column("E:E", 11)
        sheet.set_column("F:F", 38)
        sheet.set_column("G:G", 13)
        sheet.set_column("H:AM", 10)
        sheet.set_row(0, 32)
        candidate_headers = ["DISTANCE", "DEGREE", "方位", "候選排名", "原始列", "判定說明"]
        for column, header in enumerate(candidate_headers):
            sheet.write(0, column, header, formats["header"])

        for direction_index, direction in enumerate(DIRECTION_ORDER):
            column = 7 + direction_index
            if direction_index in OPEN_SEA_INDEXES:
                sheet.write(0, column, direction, formats["header"])
            else:
                sheet.write(0, column, direction, formats["excluded"])
        sheet.write(1, 6, "自動值", formats["subheader"])
        sheet.write(2, 6, "人工覆核值", formats["subheader"])
        sheet.write(3, 6, "最終值", formats["subheader"])
        sheet.write(4, 6, "判定", formats["subheader"])
        sheet.write(5, 6, "來源檔", formats["subheader"])

        candidate_rows: list[tuple[float, float, str, int, int, str]] = []
        for direction_index, direction in enumerate(DIRECTION_ORDER):
            column = 7 + direction_index
            if direction_index not in OPEN_SEA_INDEXES:
                for row in range(1, 6):
                    sheet.write(row, column, "不處理" if row == 4 else "", formats["excluded"])
                continue
            direction_result = day_result.directions[direction]
            selected = direction_result.selected
            status_format = formats["auto"] if direction_result.status == "自動採用" else formats["review"] if direction_result.status == "待複核" else formats["missing"]
            if selected is None:
                sheet.write_blank(1, column, None, status_format)
            else:
                sheet.write_number(1, column, selected, status_format)
            sheet.write_blank(2, column, None, formats["input"])
            sheet.data_validation(2, column, 2, column, {"validate": "decimal", "criteria": "between", "minimum": 0, "maximum": config.max_distance, "input_title": "人工覆核值", "input_message": "若不同意自動值，請輸入最終距離（NM）。"})
            cell = f"{_excel_col(column)}"
            final_formula = f'=IFERROR(IF(ISNUMBER({cell}3),{cell}3,IF(ISNUMBER({cell}2),{cell}2,"")),"")'
            sheet.write_formula(3, column, final_formula, formats["final"], selected if selected is not None else "")
            sheet.write(4, column, direction_result.status, status_format)
            sheet.write(5, column, day_result.source_file.name if day_result.source_file else "缺檔", formats["note"])

            for list_position, candidate in enumerate(direction_result.candidates, start=1):
                rank = candidate.rank or list_position
                note = ""
                if direction_result.selected_rank is not None:
                    if rank < direction_result.selected_rank:
                        note = "孤立高值，不採用"
                    elif direction_result.selected_rank <= rank < direction_result.selected_rank + direction_result.cluster_count:
                        note = "採用群聚"
                candidate_rows.append((candidate.distance, candidate.bearing, direction, rank, candidate.source_row, note))

        for row, values in enumerate(candidate_rows, start=1):
            sheet.write_number(row, 0, values[0], formats["number"])
            sheet.write_number(row, 1, values[1], formats["number"])
            sheet.write(row, 2, values[2], formats["body"])
            sheet.write_number(row, 3, values[3], formats["integer"])
            sheet.write_number(row, 4, values[4], formats["integer"])
            sheet.write(row, 5, values[5], formats["body"])
        if candidate_rows:
            sheet.autofilter(0, 0, len(candidate_rows), 5)
        sheet.set_landscape()
        sheet.fit_to_pages(1, 0)
        emit(callback, kind="sheet_written", position=day_index, total=len(day_results), sheet=sheet_name)

    total_sheet = workbook.add_worksheet("總表")
    total_sheet.hide_gridlines(2)
    total_sheet.freeze_panes(4, 1)
    total_sheet.set_column("A:A", 13)
    total_sheet.set_column("B:AG", 10)
    total_sheet.set_column("AI:AL", 13)
    total_sheet.set_row(0, 34)
    total_sheet.merge_range("A1:AG1", f"KLNG {config.year} 年 {config.month} 月 AIS 32方位通訊距離", formats["title"])
    total_sheet.merge_range("A2:AG2", f"自動規則：{config.max_distance:g} NM 上限；{config.tolerance:.0%} 群聚；至少 {config.cluster_size} 筆。黃色／紅色儲存格請查看「待複核」。", formats["note"])
    total_sheet.write(3, 0, "日期", formats["header"])
    for direction_index, direction in enumerate(DIRECTION_ORDER):
        column = 1 + direction_index
        total_sheet.write(3, column, direction, formats["header"] if direction_index in OPEN_SEA_INDEXES else formats["excluded"])

    first_data_row_excel = 5
    last_data_row_excel = 4 + len(day_results)
    for row_index, day_result in enumerate(day_results, start=4):
        total_sheet.write_datetime(row_index, 0, datetime.combine(day_result.day, datetime.min.time()), formats["date"])
        source_sheet = daily_sheet_names[day_result.day]
        for direction_index, direction in enumerate(DIRECTION_ORDER):
            column = 1 + direction_index
            if direction_index not in OPEN_SEA_INDEXES:
                total_sheet.write_blank(row_index, column, None, formats["excluded"])
                continue
            direction_result = day_result.directions[direction]
            source_column = _excel_col(7 + direction_index)
            formula = f'=IFERROR(\'{source_sheet}\'!{source_column}4,"")'
            cached = direction_result.selected if direction_result.selected is not None else ""
            status_format = formats["auto"] if direction_result.status == "自動採用" else formats["review"] if direction_result.status == "待複核" else formats["missing"]
            total_sheet.write_formula(row_index, column, formula, status_format, cached)

    stat_start_row = 5 + len(day_results)
    stat_specs = [("MAX", "MAX"), ("MIN", "MIN"), ("平均", "AVERAGE"), ("標準差", "STDEV.S")]
    stat_values_by_direction: dict[str, tuple[float | None, float | None, float | None, float | None]] = {}
    for direction in DIRECTION_ORDER:
        stat_values_by_direction[direction] = _stats(day.directions[direction].selected for day in day_results)
    for stat_offset, (label, function) in enumerate(stat_specs):
        row = stat_start_row + stat_offset
        total_sheet.write(row, 0, label, formats["stat_label"])
        for direction_index, direction in enumerate(DIRECTION_ORDER):
            column = 1 + direction_index
            if direction_index not in OPEN_SEA_INDEXES:
                total_sheet.write_blank(row, column, None, formats["excluded"])
                continue
            col_letter = _excel_col(column)
            formula = f'=IFERROR({function}({col_letter}{first_data_row_excel}:{col_letter}{last_data_row_excel}),"")'
            cached = stat_values_by_direction[direction][stat_offset]
            total_sheet.write_formula(row, column, formula, formats["stat"], cached if cached is not None else "")

    helper_row = 3
    total_sheet.write(helper_row, 34, "海向方位", formats["header"])
    total_sheet.write(helper_row, 35, "MAX", formats["header"])
    total_sheet.write(helper_row, 36, "MIN", formats["header"])
    total_sheet.write(helper_row, 37, "平均", formats["header"])
    for helper_offset, direction_index in enumerate(OPEN_SEA_INDEXES, start=1):
        row = helper_row + helper_offset
        direction = DIRECTION_ORDER[direction_index]
        source_col = _excel_col(1 + direction_index)
        total_sheet.write(row, 34, direction, formats["body"])
        for stat_offset, helper_col in enumerate((35, 36, 37)):
            source_row_excel = stat_start_row + stat_offset + 1
            formula = f'=IFERROR({source_col}{source_row_excel},"")'
            cached = stat_values_by_direction[direction][stat_offset]
            total_sheet.write_formula(row, helper_col, formula, formats["number"], cached if cached is not None else "")

    radar = workbook.add_chart({"type": "radar", "subtype": "with_markers"})
    palette = ["#2F6B7C", "#D97706", "#7C3AED", "#15803D", "#B91C1C", "#0369A1", "#A21CAF"]
    for day_offset, day_result in enumerate(day_results):
        if not any(day_result.directions[DIRECTION_ORDER[index]].selected is not None for index in OPEN_SEA_INDEXES):
            continue
        excel_row = first_data_row_excel + day_offset
        radar.add_series(
            {
                "name": f"{config.month}/{day_result.day.day}",
                "categories": ["總表", 3, 1, 3, 32],
                "values": ["總表", excel_row - 1, 1, excel_row - 1, 32],
                "line": {"color": palette[day_offset % len(palette)], "width": 0.75, "transparency": 45},
                "marker": {"type": "none"},
            }
        )
    radar.set_title({"name": f"{config.year}-{config.month:02d} AIS 通訊涵蓋圖"})
    radar.set_legend({"position": "bottom", "font": {"size": 8}})
    radar.set_size({"width": 1080, "height": 560})
    radar.set_style(10)

    columns = workbook.add_chart({"type": "column"})
    category_range = ["總表", helper_row + 1, 34, helper_row + len(OPEN_SEA_INDEXES), 34]
    for series_name, helper_col, color in (("最遠", 35, "#2F6B7C"), ("最近", 36, "#D97706")):
        columns.add_series(
            {
                "name": series_name,
                "categories": category_range,
                "values": ["總表", helper_row + 1, helper_col, helper_row + len(OPEN_SEA_INDEXES), helper_col],
                "fill": {"color": color, "transparency": 10},
                "border": {"none": True},
            }
        )
    line = workbook.add_chart({"type": "line"})
    line.add_series(
        {
            "name": "平均",
            "categories": category_range,
            "values": ["總表", helper_row + 1, 37, helper_row + len(OPEN_SEA_INDEXES), 37],
            "line": {"color": "#B91C1C", "width": 2.25},
            "marker": {"type": "circle", "size": 4, "border": {"color": "#B91C1C"}, "fill": {"color": "#FFFFFF"}},
        }
    )
    columns.combine(line)
    columns.set_title({"name": f"{config.year}-{config.month:02d} AIS 通訊最近、最遠與平均距離（NM）"})
    columns.set_y_axis({"name": "海里（NM）", "major_gridlines": {"visible": True, "line": {"color": "#D9E2E8"}}, "num_format": "0"})
    columns.set_x_axis({"label_position": "low", "num_font": {"rotation": -45, "size": 8}})
    columns.set_legend({"position": "bottom"})
    columns.set_size({"width": 1080, "height": 520})
    columns.set_style(10)

    chart_row = stat_start_row + len(stat_specs) + 2
    total_sheet.insert_chart(chart_row, 0, radar)
    total_sheet.insert_chart(chart_row + 29, 0, columns)
    total_sheet.set_landscape()
    total_sheet.fit_to_pages(1, 0)
    total_sheet.print_area(0, 0, chart_row + 55, 32)

    review_sheet = workbook.add_worksheet("待複核")
    review_sheet.hide_gridlines(2)
    review_sheet.freeze_panes(1, 0)
    review_sheet.set_column("A:A", 12)
    review_sheet.set_column("B:B", 11)
    review_sheet.set_column("C:C", 12)
    review_sheet.set_column("D:D", 12)
    review_sheet.set_column("E:E", 12)
    review_sheet.set_column("F:F", 54)
    review_sheet.set_column("G:G", 54)
    review_sheet.set_column("H:H", 22)
    review_headers = ["日期", "方位", "自動值", "群聚筆數", "狀態", "原因", "前10筆候選距離", "人工覆核位置"]
    for column, header in enumerate(review_headers):
        review_sheet.write(0, column, header, formats["header"])
    review_row = 1
    for day_result in day_results:
        for direction_index in OPEN_SEA_INDEXES:
            direction = DIRECTION_ORDER[direction_index]
            direction_result = day_result.directions[direction]
            if direction_result.status == "自動採用":
                continue
            review_sheet.write_datetime(review_row, 0, datetime.combine(day_result.day, datetime.min.time()), formats["date"])
            review_sheet.write(review_row, 1, direction, formats["body"])
            if direction_result.selected is None:
                review_sheet.write_blank(review_row, 2, None, formats["missing"])
            else:
                review_sheet.write_number(review_row, 2, direction_result.selected, formats["review"])
            review_sheet.write_number(review_row, 3, direction_result.cluster_count, formats["integer"])
            review_sheet.write(review_row, 4, direction_result.status, formats["review"] if direction_result.status == "待複核" else formats["missing"])
            review_sheet.write(review_row, 5, direction_result.reason, formats["body_wrap"])
            review_sheet.write(review_row, 6, ", ".join(f"{candidate.distance:.3f}" for candidate in direction_result.candidates[:10]), formats["body_wrap"])
            source_sheet = daily_sheet_names[day_result.day]
            source_column = _excel_col(7 + direction_index)
            review_sheet.write_url(review_row, 7, f"internal:'{source_sheet}'!{source_column}3", formats["link"], f"{source_sheet}!{source_column}3")
            review_row += 1
    if review_row == 1:
        review_sheet.write(1, 0, "沒有待複核項目", formats["auto"])
    else:
        review_sheet.autofilter(0, 0, review_row - 1, len(review_headers) - 1)

    log_sheet = workbook.add_worksheet("處理紀錄")
    log_sheet.hide_gridlines(2)
    log_sheet.freeze_panes(1, 0)
    log_sheet.set_column("A:A", 12)
    log_sheet.set_column("B:B", 38)
    log_sheet.set_column("C:G", 14)
    log_sheet.set_column("H:H", 12)
    log_headers = ["日期", "來源檔", "掃描列數", "合格列數", "錯誤列數", "訊息類型略過", "非 East 略過", "秒數"]
    for column, header in enumerate(log_headers):
        log_sheet.write(0, column, header, formats["header"])
    for row, day_result in enumerate(day_results, start=1):
        log_sheet.write_datetime(row, 0, datetime.combine(day_result.day, datetime.min.time()), formats["date"])
        log_sheet.write(row, 1, day_result.source_file.name if day_result.source_file else "缺檔", formats["body"])
        for column, value in enumerate(
            (
                day_result.rows_scanned,
                day_result.rows_accepted,
                day_result.rows_invalid,
                day_result.rows_wrong_message,
                day_result.rows_not_east,
            ),
            start=2,
        ):
            log_sheet.write_number(row, column, value, formats["integer"])
        log_sheet.write_number(row, 7, day_result.elapsed_seconds, formats["number"])
    log_sheet.autofilter(0, 0, len(day_results), len(log_headers) - 1)

    try:
        workbook.close()
        os.replace(temp_path, config.output_path)
    except Exception:
        try:
            workbook.close()
        except Exception:
            pass
        raise
    emit(callback, kind="write_done", file=config.output_path.name)
    return config.output_path


def write_legacy_workbook(
    config: AppConfig,
    day_results: list[DayResult],
    callback: ProgressCallback | None = None,
    cancel_event: threading.Event | None = None,
) -> Path | None:
    output_path = config.legacy_output_path
    if output_path is None:
        return None
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(output_path.stem + ".building.xlsx")
    if temp_path.exists():
        temp_path.unlink()

    emit(callback, kind="legacy_write_start", file=output_path.name)
    workbook = xlsxwriter.Workbook(
        temp_path,
        {
            "constant_memory": True,
            "nan_inf_to_errors": True,
            "default_date_format": "m月d日",
        },
    )
    workbook.use_zip64()
    workbook.set_properties(
        {
            "title": f"KLNG {config.year}年{config.month}月 32方位數值 原格式相容版",
            "subject": "完整排序資料、32方位結果與舊式總表",
            "author": "AIS 32方位月報一鍵製作",
            "comments": "A:C 保留舊流程的完整排序資料；H:AM 第 2 列為數值化規則自動結果。",
        }
    )
    workbook.set_calc_mode("auto")
    header_format = workbook.add_format({"bold": True, "align": "center", "bottom": 1})
    number_format = workbook.add_format({"num_format": "0.000"})
    date_format = workbook.add_format({"num_format": "m月d日"})
    stat_label_format = workbook.add_format({"bold": True})

    daily_names: dict[date, str] = {}
    try:
        for position, day_result in enumerate(day_results, start=1):
            if cancel_event and cancel_event.is_set():
                raise CancelledError("使用者已取消")
            sheet_name = _safe_sheet_name(f"{config.month}月{day_result.day.day}日")
            daily_names[day_result.day] = sheet_name
            sheet = workbook.add_worksheet(sheet_name)
            sheet.freeze_panes(1, 0)
            sheet.set_column("A:A", 14)
            sheet.set_column("B:B", 12)
            sheet.set_column("C:C", 11)
            sheet.set_column("D:G", 3)
            sheet.set_column("H:AM", 11)
            sheet.write_row(0, 0, ["DISTANCE", "DEGREE", "方位"], header_format)
            sheet.write_row(0, 7, DIRECTION_ORDER, header_format)
            for direction_index, direction in enumerate(DIRECTION_ORDER):
                selected = day_result.directions[direction].selected
                if direction_index in OPEN_SEA_INDEXES and selected is not None:
                    sheet.write_number(1, 7 + direction_index, selected, number_format)

            spool_path = _legacy_spool_path(config, day_result.day)
            if day_result.source_file is not None:
                if not spool_path.is_file():
                    raise FileNotFoundError(
                        f"缺少 {day_result.day:%Y-%m-%d} 的原格式資料暫存檔；請重新執行該月份。"
                    )
                counts, records = read_legacy_spool(spool_path)
                if sum(counts) != day_result.rows_legacy:
                    raise ValueError(
                        f"{day_result.day:%Y-%m-%d} 原格式資料筆數不一致："
                        f"暫存 {sum(counts):,} / 紀錄 {day_result.rows_legacy:,}。"
                    )
                excel_row = 1
                for direction_index, distance, bearing in records:
                    if cancel_event and excel_row % 10_000 == 0 and cancel_event.is_set():
                        raise CancelledError("使用者已取消")
                    sheet.write_number(excel_row, 0, distance)
                    sheet.write_number(excel_row, 1, bearing)
                    sheet.write(excel_row, 2, DIRECTION_ORDER[direction_index])
                    excel_row += 1
            emit(
                callback,
                kind="legacy_sheet_written",
                position=position,
                total=len(day_results),
                sheet=sheet_name,
                rows=day_result.rows_legacy,
            )

        summary = workbook.add_worksheet("總表")
        summary.freeze_panes(1, 1)
        summary.set_column("A:A", 11)
        summary.set_column("B:AG", 10)
        summary.write_row(0, 1, LEGACY_DIRECTION_ABBREVIATIONS, header_format)
        for row, day_result in enumerate(day_results, start=1):
            summary.write_datetime(row, 0, datetime.combine(day_result.day, datetime.min.time()), date_format)
            source_sheet = daily_names[day_result.day]
            for direction_index, direction in enumerate(DIRECTION_ORDER):
                if direction_index not in OPEN_SEA_INDEXES:
                    summary.write_blank(row, 1 + direction_index, None, number_format)
                    continue
                source_column = _excel_col(7 + direction_index)
                selected = day_result.directions[direction].selected
                formula = f"=IFERROR('{source_sheet}'!{source_column}2,\"\")"
                summary.write_formula(row, 1 + direction_index, formula, number_format, selected if selected is not None else "")

        stat_start = len(day_results) + 2
        stat_specs = [("MAX", "MAX", 0), ("MIN", "MIN", 1), ("平均值", "AVERAGE", 2)]
        stats_by_direction = {
            direction: (
                _stats(day.directions[direction].selected for day in day_results)
                if direction_index in OPEN_SEA_INDEXES
                else (None, None, None, None)
            )
            for direction_index, direction in enumerate(DIRECTION_ORDER)
        }
        for offset, (label, _function, stat_index) in enumerate(stat_specs):
            row = stat_start + offset
            summary.write(row, 0, label, stat_label_format)
            for direction_index, direction in enumerate(DIRECTION_ORDER):
                column = 1 + direction_index
                if direction_index not in OPEN_SEA_INDEXES:
                    summary.write_blank(row, column, None, number_format)
                    continue
                cached = stats_by_direction[direction][stat_index]
                if cached is None:
                    summary.write_blank(row, column, None, number_format)
                else:
                    summary.write_number(row, column, cached, number_format)

        radar = workbook.add_chart({"type": "radar"})
        palette = ["#4472C4", "#ED7D31", "#70AD47", "#A5A5A5", "#FFC000", "#5B9BD5"]
        for day_offset, day_result in enumerate(day_results):
            if not any(
                day_result.directions[DIRECTION_ORDER[index]].selected is not None
                for index in OPEN_SEA_INDEXES
            ):
                continue
            row = 1 + day_offset
            radar.add_series(
                {
                    "name": f"{config.month}月{day_result.day.day}日",
                    "categories": ["總表", 0, 1, 0, 32],
                    "values": ["總表", row, 1, row, 32],
                    "line": {"color": palette[day_offset % len(palette)], "width": 0.75, "transparency": 35},
                }
            )
        radar.set_title({"name": f"{config.year}-{config.month:02d} AIS通訊涵蓋圖"})
        radar.set_legend({"position": "right", "font": {"size": 8}})
        radar.set_size({"width": 760, "height": 480})
        summary.insert_chart(stat_start + 6, 0, radar)

        columns = workbook.add_chart({"type": "column"})
        categories = ["總表", 0, 1, 0, 32]
        for label, row, color in (
            ("最遠距離", stat_start, "#4472C4"),
            ("最近距離", stat_start + 1, "#ED7D31"),
            ("平均值", stat_start + 2, "#70AD47"),
        ):
            columns.add_series(
                {
                    "name": label,
                    "categories": categories,
                    "values": ["總表", row, 1, row, 32],
                    "fill": {"color": color},
                    "border": {"none": True},
                }
            )
        columns.set_title({"name": f"{config.year}-{config.month:02d} AIS通訊最遠、最近與平均距離（NM）"})
        columns.set_legend({"position": "bottom"})
        columns.set_y_axis({"name": "NM", "num_format": "0"})
        columns.set_size({"width": 760, "height": 480})
        summary.insert_chart(stat_start + 31, 0, columns)

        mapping = workbook.add_worksheet("工作")
        mapping.write_row(0, 9, ["POSITION", "方位"], header_format)
        for degree in range(361):
            direction_info = degree_to_direction_index(degree)
            direction = DIRECTION_ORDER[direction_info[0]] if direction_info else ""
            mapping.write_number(degree + 1, 9, degree)
            mapping.write(degree + 1, 10, direction)
        mapping.hide()

        workbook.close()
        os.replace(temp_path, output_path)
    except Exception:
        try:
            workbook.close()
        except Exception:
            pass
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise
    emit(callback, kind="legacy_write_done", file=output_path.name)
    return output_path


def run_pipeline(
    config: AppConfig,
    callback: ProgressCallback | None = None,
    cancel_event: threading.Event | None = None,
) -> Path:
    started = time.perf_counter()
    day_results, warnings = process_month(config, callback=callback, cancel_event=cancel_event)
    output = write_monthly_workbook(config, day_results, warnings, callback=callback, cancel_event=cancel_event)
    legacy_output = write_legacy_workbook(config, day_results, callback=callback, cancel_event=cancel_event)
    files = [str(output)] + ([str(legacy_output)] if legacy_output is not None else [])
    emit(callback, kind="complete", file=str(output), files=files, seconds=time.perf_counter() - started)
    return output


def parse_message_types(text: str) -> tuple[int, ...]:
    values: list[int] = []
    for item in re.split(r"[,，\s]+", text.strip()):
        if not item:
            continue
        value = int(item)
        if value not in values:
            values.append(value)
    return tuple(values)


def format_seconds(seconds: float) -> str:
    seconds_int = max(int(round(seconds)), 0)
    hours, remainder = divmod(seconds_int, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours} 小時 {minutes} 分 {secs} 秒"
    if minutes:
        return f"{minutes} 分 {secs} 秒"
    return f"{secs} 秒"


def derive_legacy_output_path(modern_output: Path) -> Path:
    stem = modern_output.stem
    for suffix in ("_新版自動分析", "_自動分析版", "_新版"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
            break
    return modern_output.with_name(f"{stem}_原格式相容版.xlsx")


def friendly_error_message(error: BaseException) -> str:
    if isinstance(error, CancelledError):
        return "已依使用者要求安全停止；正式輸出檔不會被半成品覆寫。"
    if isinstance(error, SourceFileError):
        return str(error)
    if isinstance(error, FileExistsError):
        return f"輸出檔已存在。請改檔名、移走舊檔，或確認允許覆寫。\n\n{error}"
    if isinstance(error, PermissionError):
        return (
            "無法讀寫檔案。最常見原因是輸出 Excel 正在被其他人開啟，或資料夾沒有寫入權限。"
            "請關閉相關 Excel、確認共用磁碟連線後再試。\n\n"
            f"原始錯誤：{error}"
        )
    if isinstance(error, (zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException)):
        return f"其中一個來源檔不是有效的 Excel .xlsx，可能下載未完成或檔案損壞。\n\n原始錯誤：{error}"
    if isinstance(error, MemoryError):
        return "記憶體不足。請把「平行檔案數」降為 1 或 2，關閉其他大型程式後重跑；已完成日期可由快取續跑。"
    if isinstance(error, OSError):
        error_number = getattr(error, "errno", None)
        windows_error = getattr(error, "winerror", None)
        if error_number == 28 or windows_error == 112:
            return "磁碟空間不足。請清出至少 2 GB，或將輸出位置改到空間較大的磁碟後重跑。"
        return f"檔案系統發生錯誤。請確認來源／輸出磁碟仍連線且有足夠空間。\n\n原始錯誤：{error}"
    if isinstance(error, ValueError):
        return str(error)
    if type(error).__name__ in {"BrokenProcessPool", "BrokenExecutor"}:
        return "平行處理程序異常停止。請將「平行檔案數」降為 1 或 2 後重跑；已完成日期會沿用快取。"
    return (
        "發生未預期錯誤。程式已保留已完成日期的快取；請依錯誤報告中的檔名與訊息處理後重跑。\n\n"
        f"{type(error).__name__}: {error}"
    )


def write_error_report(config: AppConfig, error: BaseException) -> Path | None:
    try:
        config.output_path.parent.mkdir(parents=True, exist_ok=True)
        report = config.output_path.parent / f"AIS月報_錯誤報告_{datetime.now():%Y%m%d_%H%M%S}.txt"
        contents = [
            APP_TITLE,
            f"版本：{APP_VERSION}",
            f"時間：{datetime.now():%Y-%m-%d %H:%M:%S}",
            f"來源：{config.input_dir}",
            f"月份：{config.year}-{config.month:02d}",
            f"新版輸出：{config.output_path}",
            f"原格式輸出：{config.legacy_output_path or '未要求'}",
            "",
            "給使用者的說明：",
            friendly_error_message(error),
            "",
            "技術細節：",
            "".join(traceback.format_exception(type(error), error, error.__traceback__)),
        ]
        report.write_text("\n".join(contents), encoding="utf-8-sig")
        return report
    except OSError:
        return None


def launch_gui() -> None:
    # PyInstaller one-file extracts Tcl/Tk beside the frozen program.  Set the
    # paths explicitly before importing tkinter because some Windows Python
    # distributions fail to apply the bundled runtime hook early enough.
    bundle_root = getattr(sys, "_MEIPASS", None)
    if bundle_root:
        bundle_path = Path(bundle_root)
        os.environ["TCL_LIBRARY"] = str(bundle_path / "_tcl_data")
        os.environ["TK_LIBRARY"] = str(bundle_path / "_tk_data")
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    class Application(tk.Tk):
        def __init__(self) -> None:
            super().__init__()
            self.title(f"{APP_TITLE} v{APP_VERSION}")
            self.geometry("960x760")
            self.minsize(860, 700)
            self.configure(bg="#F4F7F9")
            self.events: queue.Queue[dict] = queue.Queue()
            self.cancel_event = threading.Event()
            self.worker: threading.Thread | None = None
            self.last_output: Path | None = None
            self.last_outputs: list[Path] = []

            self.source_var = tk.StringVar()
            self.output_var = tk.StringVar()
            self.legacy_var = tk.BooleanVar(value=True)
            self.legacy_path_var = tk.StringVar(value="將依新版檔名自動建立")
            self.year_var = tk.StringVar(value=str(datetime.now().year))
            self.month_var = tk.StringVar(value=str(datetime.now().month))
            self.max_distance_var = tk.StringVar(value="500")
            self.tolerance_var = tk.StringVar(value="10")
            self.cluster_var = tk.StringVar(value="3")
            self.top_var = tk.StringVar(value="50")
            self.workers_var = tk.StringVar(value=str(min(3, os.cpu_count() or 1)))
            self.message_types_var = tk.StringVar(value="1,2,3,18,19")
            self.status_var = tk.StringVar(value="請先選擇每日 AIS Excel 所在資料夾。")
            self.progress_text_var = tk.StringVar(value="尚未開始")

            style = ttk.Style(self)
            style.theme_use("vista")
            style.configure("Title.TLabel", background="#17324D", foreground="white", font=("Microsoft JhengHei", 18, "bold"), padding=(18, 14))
            style.configure("Section.TLabelframe", background="#F4F7F9", padding=12)
            style.configure("Section.TLabelframe.Label", background="#F4F7F9", foreground="#17324D", font=("Microsoft JhengHei", 11, "bold"))
            style.configure("Primary.TButton", font=("Microsoft JhengHei", 12, "bold"), padding=(18, 10))
            style.configure("TLabel", background="#F4F7F9", font=("Microsoft JhengHei", 10))
            style.configure("TButton", font=("Microsoft JhengHei", 10))

            ttk.Label(self, text=APP_TITLE, style="Title.TLabel").pack(fill="x")
            container = ttk.Frame(self, padding=16)
            container.pack(fill="both", expand=True)
            container.columnconfigure(0, weight=1)

            files_box = ttk.LabelFrame(container, text="1  檔案位置", style="Section.TLabelframe")
            files_box.grid(row=0, column=0, sticky="ew", pady=(0, 10))
            files_box.columnconfigure(1, weight=1)
            ttk.Label(files_box, text="每日資料夾").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=5)
            ttk.Entry(files_box, textvariable=self.source_var).grid(row=0, column=1, sticky="ew", pady=5)
            ttk.Button(files_box, text="選擇…", command=self.choose_source).grid(row=0, column=2, padx=(8, 0), pady=5)
            ttk.Label(files_box, text="輸出月報").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=5)
            ttk.Entry(files_box, textvariable=self.output_var).grid(row=1, column=1, sticky="ew", pady=5)
            ttk.Button(files_box, text="另存為…", command=self.choose_output).grid(row=1, column=2, padx=(8, 0), pady=5)
            ttk.Checkbutton(files_box, text="同時產生原格式相容版", variable=self.legacy_var, command=self.update_legacy_path).grid(row=2, column=0, sticky="w", pady=5)
            ttk.Label(files_box, textvariable=self.legacy_path_var, foreground="#5B6573").grid(row=2, column=1, columnspan=2, sticky="w", pady=5)
            self.output_var.trace_add("write", lambda *_args: self.update_legacy_path())

            config_box = ttk.LabelFrame(container, text="2  月份與判斷設定", style="Section.TLabelframe")
            config_box.grid(row=1, column=0, sticky="ew", pady=(0, 10))
            labels = [
                ("年份", self.year_var, 8),
                ("月份", self.month_var, 5),
                ("距離上限 NM", self.max_distance_var, 8),
                ("群聚差距 %", self.tolerance_var, 7),
                ("至少筆數", self.cluster_var, 5),
                ("每方向候選", self.top_var, 6),
                ("平行檔案數", self.workers_var, 5),
            ]
            for column, (label, variable, width) in enumerate(labels):
                ttk.Label(config_box, text=label).grid(row=0, column=column, sticky="w", padx=(0 if column == 0 else 10, 0))
                ttk.Entry(config_box, textvariable=variable, width=width).grid(row=1, column=column, sticky="w", padx=(0 if column == 0 else 10, 0), pady=(3, 8))
            ttk.Label(config_box, text="AIS 訊息類型").grid(row=2, column=0, sticky="w")
            ttk.Entry(config_box, textvariable=self.message_types_var, width=24).grid(row=2, column=1, columnspan=2, sticky="w", padx=(10, 0))
            ttk.Label(config_box, text="船舶位置預設 1,2,3,18,19；每日完成即保存快取，可中斷後續跑。", foreground="#5B6573").grid(row=2, column=3, columnspan=4, sticky="w", padx=(12, 0))

            run_box = ttk.LabelFrame(container, text="3  一鍵執行", style="Section.TLabelframe")
            run_box.grid(row=2, column=0, sticky="nsew")
            run_box.columnconfigure(0, weight=1)
            container.rowconfigure(2, weight=1)
            button_row = ttk.Frame(run_box)
            button_row.grid(row=0, column=0, sticky="ew")
            self.start_button = ttk.Button(button_row, text="開始全自動製作", style="Primary.TButton", command=self.start)
            self.start_button.pack(side="left")
            self.cancel_button = ttk.Button(button_row, text="取消", command=self.cancel, state="disabled")
            self.cancel_button.pack(side="left", padx=8)
            self.open_button = ttk.Button(button_row, text="開啟輸出資料夾", command=self.open_output, state="disabled")
            self.open_button.pack(side="right")
            ttk.Label(run_box, textvariable=self.status_var, foreground="#17324D").grid(row=1, column=0, sticky="w", pady=(12, 3))
            self.progress = ttk.Progressbar(run_box, mode="determinate", maximum=100)
            self.progress.grid(row=2, column=0, sticky="ew", pady=3)
            ttk.Label(run_box, textvariable=self.progress_text_var, foreground="#5B6573").grid(row=3, column=0, sticky="w", pady=(0, 8))
            self.log = tk.Text(run_box, height=14, wrap="word", font=("Consolas", 9), bg="#FFFFFF", fg="#23313D", relief="solid", borderwidth=1)
            self.log.grid(row=4, column=0, sticky="nsew")
            run_box.rowconfigure(4, weight=1)
            scrollbar = ttk.Scrollbar(run_box, orient="vertical", command=self.log.yview)
            scrollbar.grid(row=4, column=1, sticky="ns")
            self.log.configure(yscrollcommand=scrollbar.set, state="disabled")
            self.after(150, self.poll_events)

        def append_log(self, text: str) -> None:
            self.log.configure(state="normal")
            self.log.insert("end", f"[{datetime.now():%H:%M:%S}] {text}\n")
            self.log.see("end")
            self.log.configure(state="disabled")

        def choose_source(self) -> None:
            selected = filedialog.askdirectory(title="選擇 D&TMOK KLNG 每日 Excel 資料夾")
            if not selected:
                return
            source = Path(selected)
            self.source_var.set(str(source))
            months = detect_available_months(source)
            if months:
                year, month, count = max(months, key=lambda item: (item[2], item[0], item[1]))
                self.year_var.set(str(year))
                self.month_var.set(str(month))
                default_dir = Path(__file__).resolve().parent.parent / "output" / "AIS月報"
                default_name = f"KLNG {year}年{month}月 32方位數值_新版自動分析.xlsx"
                self.output_var.set(str(default_dir / default_name))
                self.status_var.set(f"偵測到 {year} 年 {month} 月，共 {count} 個每日檔。")
                self.append_log(self.status_var.get())
            else:
                self.status_var.set("找不到符合 D&TMOK KLNG_YYYYMMDD_*.xlsx 格式的檔案。")

        def choose_output(self) -> None:
            initial = Path(self.output_var.get()) if self.output_var.get() else Path.cwd() / "KLNG 32方位數值.xlsx"
            selected = filedialog.asksaveasfilename(
                title="儲存 AIS 月報",
                initialdir=str(initial.parent),
                initialfile=initial.name,
                defaultextension=".xlsx",
                filetypes=[("Excel 活頁簿", "*.xlsx")],
            )
            if selected:
                self.output_var.set(selected)

        def update_legacy_path(self) -> None:
            raw = self.output_var.get().strip()
            if not self.legacy_var.get():
                self.legacy_path_var.set("不產生原格式相容版")
            elif raw:
                self.legacy_path_var.set(str(derive_legacy_output_path(Path(raw))))
            else:
                self.legacy_path_var.set("將依新版檔名自動建立")

        def build_config(self) -> AppConfig:
            source = Path(self.source_var.get().strip())
            output = Path(self.output_var.get().strip())
            if not str(source):
                raise ValueError("請選擇每日資料夾。")
            if not str(output):
                raise ValueError("請指定輸出月報。")
            legacy_output = derive_legacy_output_path(output) if self.legacy_var.get() else None
            return AppConfig(
                input_dir=source,
                output_path=output,
                year=int(self.year_var.get()),
                month=int(self.month_var.get()),
                legacy_output_path=legacy_output,
                max_distance=float(self.max_distance_var.get()),
                tolerance=float(self.tolerance_var.get()) / 100.0,
                cluster_size=int(self.cluster_var.get()),
                top_candidates=int(self.top_var.get()),
                message_types=parse_message_types(self.message_types_var.get()),
                workers=int(self.workers_var.get()),
                overwrite=True,
            )

        def start(self) -> None:
            try:
                config = self.build_config()
                existing = [path for path in (config.output_path, config.legacy_output_path) if path is not None and path.exists()]
                if existing:
                    listing = "\n".join(str(path) for path in existing)
                    if not messagebox.askyesno("覆寫確認", f"下列檔案已存在：\n{listing}\n\n要覆寫嗎？"):
                        return
                config.validate()
            except Exception as error:
                messagebox.showerror("設定錯誤", str(error))
                return
            self.cancel_event.clear()
            self.last_output = None
            self.last_outputs = []
            self.start_button.configure(state="disabled")
            self.cancel_button.configure(state="normal")
            self.open_button.configure(state="disabled")
            self.progress["value"] = 0
            self.status_var.set("開始處理…")
            self.progress_text_var.set("正在準備來源檔")
            self.append_log(f"開始：{config.year} 年 {config.month} 月")
            self.append_log(f"新版：{config.output_path}")
            if config.legacy_output_path is not None:
                self.append_log(f"原格式版：{config.legacy_output_path}")

            def callback(payload: dict) -> None:
                self.events.put(payload)

            def work() -> None:
                try:
                    output = run_pipeline(config, callback=callback, cancel_event=self.cancel_event)
                    outputs = [str(output)]
                    if config.legacy_output_path is not None:
                        outputs.append(str(config.legacy_output_path))
                    self.events.put({"kind": "worker_success", "output": str(output), "outputs": outputs})
                except CancelledError as error:
                    self.events.put({"kind": "worker_cancelled", "error": str(error)})
                except Exception as error:
                    report = write_error_report(config, error)
                    self.events.put(
                        {
                            "kind": "worker_error",
                            "error": friendly_error_message(error),
                            "technical": f"{type(error).__name__}: {error}",
                            "report": str(report) if report else None,
                        }
                    )

            self.worker = threading.Thread(target=work, daemon=True)
            self.worker.start()

        def cancel(self) -> None:
            self.cancel_event.set()
            self.cancel_button.configure(state="disabled")
            self.status_var.set("正在安全停止；目前來源檔讀完或每 10,000 列會檢查一次…")
            self.append_log("已要求取消。")

        def poll_events(self) -> None:
            try:
                while True:
                    payload = self.events.get_nowait()
                    kind = payload.get("kind")
                    if kind == "file_start":
                        position, total = int(payload["position"]), int(payload["total"])
                        self.progress["value"] = (position - 1) / max(total, 1) * 85
                        self.status_var.set(f"處理第 {position}/{total} 檔：{payload['file']}")
                        self.progress_text_var.set("正在逐列讀取並保留各方位候選值")
                        self.append_log(self.status_var.get())
                    elif kind == "preflight":
                        required = int(payload["required_bytes"]) / 1024**3
                        free_value = payload.get("free_bytes")
                        free_text = f"，可用 {int(free_value) / 1024**3:.1f} GB" if free_value else ""
                        self.append_log(f"輸出空間預估至少 {required:.1f} GB{free_text}")
                    elif kind == "parallel_start":
                        self.status_var.set(f"使用 {payload['workers']} 個處理程序平行讀取 {payload['total']} 個來源檔…")
                        self.append_log(self.status_var.get())
                    elif kind == "cache_hit":
                        position, total = int(payload["position"]), int(payload["total"])
                        self.progress["value"] = position / max(total, 1) * 85
                        self.progress_text_var.set(f"沿用已完成快取：{payload['file']}")
                        self.append_log(f"快取命中：{payload['file']}")
                    elif kind == "rows":
                        rows, estimated = int(payload["rows"]), int(payload["estimated_rows"])
                        self.progress_text_var.set(f"{payload['file']}：已掃描 {rows:,} / 約 {estimated:,} 列")
                    elif kind == "file_done":
                        position, total = int(payload["position"]), int(payload["total"])
                        self.progress["value"] = position / max(total, 1) * 85
                        self.append_log(f"完成 {payload['file']}，耗時 {format_seconds(float(payload['seconds']))}")
                    elif kind == "write_start":
                        self.progress["value"] = 88
                        self.status_var.set("資料讀取完成，正在製作 Excel 月報與圖表…")
                        self.progress_text_var.set(payload["file"])
                        self.append_log(self.status_var.get())
                    elif kind == "sheet_written":
                        position, total = int(payload["position"]), int(payload["total"])
                        self.progress["value"] = 88 + position / max(total, 1) * 10
                        self.progress_text_var.set(f"已建立 {payload['sheet']}（{position}/{total}）")
                    elif kind == "write_done":
                        self.progress["value"] = 92 if self.legacy_var.get() else 99
                        self.append_log(f"Excel 已寫入：{payload['file']}")
                    elif kind == "legacy_write_start":
                        self.progress["value"] = 92
                        self.status_var.set("新版完成，正在製作原格式完整資料版…")
                        self.progress_text_var.set(payload["file"])
                        self.append_log(self.status_var.get())
                    elif kind == "legacy_sheet_written":
                        position, total = int(payload["position"]), int(payload["total"])
                        self.progress["value"] = 92 + position / max(total, 1) * 7
                        self.progress_text_var.set(
                            f"原格式：{payload['sheet']}，寫入 {int(payload['rows']):,} 列（{position}/{total}）"
                        )
                    elif kind == "legacy_write_done":
                        self.progress["value"] = 99
                        self.append_log(f"原格式 Excel 已寫入：{payload['file']}")
                    elif kind == "complete":
                        self.progress_text_var.set(f"總耗時：{format_seconds(float(payload['seconds']))}")
                    elif kind == "worker_success":
                        self.last_output = Path(payload["output"])
                        self.last_outputs = [Path(value) for value in payload.get("outputs", [payload["output"]])]
                        self.progress["value"] = 100
                        self.status_var.set("完成。新版分析版與原格式相容版均已產生。" if len(self.last_outputs) == 2 else "完成。新版分析月報已產生。")
                        for completed_output in self.last_outputs:
                            self.append_log(f"完成：{completed_output}")
                        self.start_button.configure(state="normal")
                        self.cancel_button.configure(state="disabled")
                        self.open_button.configure(state="normal")
                        listing = "\n".join(str(path) for path in self.last_outputs)
                        messagebox.showinfo("製作完成", f"AIS 月報已完成：\n{listing}\n\n請先查看新版的「待複核」工作表。")
                    elif kind == "worker_cancelled":
                        self.status_var.set("已取消；未覆寫正式輸出檔。")
                        self.append_log(payload["error"])
                        self.start_button.configure(state="normal")
                        self.cancel_button.configure(state="disabled")
                    elif kind == "worker_error":
                        self.status_var.set("製作失敗；請看下方錯誤。")
                        self.append_log(payload.get("technical", payload["error"]))
                        if payload.get("report"):
                            self.append_log(f"錯誤報告：{payload['report']}")
                        self.start_button.configure(state="normal")
                        self.cancel_button.configure(state="disabled")
                        report_note = f"\n\n完整錯誤報告：\n{payload['report']}" if payload.get("report") else ""
                        messagebox.showerror("製作失敗", payload["error"] + report_note)
            except queue.Empty:
                pass
            self.after(150, self.poll_events)

        def open_output(self) -> None:
            if self.last_output:
                subprocess.Popen(["explorer", "/select,", str(self.last_output)])

    Application().mainloop()


def cli_progress(payload: dict) -> None:
    if sys.stdout is None:
        return
    kind = payload.get("kind")
    if kind == "file_start":
        print(f"[{payload['position']}/{payload['total']}] {payload['file']}", flush=True)
    elif kind == "rows":
        print(f"  scanned {int(payload['rows']):,} rows", flush=True)
    elif kind == "file_done":
        print(f"  done in {format_seconds(float(payload['seconds']))}", flush=True)
    elif kind == "write_start":
        print("Writing workbook and charts…", flush=True)
    elif kind == "legacy_write_start":
        print("Writing legacy-compatible full workbook…", flush=True)
    elif kind == "legacy_sheet_written":
        print(
            f"  legacy {payload['position']}/{payload['total']} {payload['sheet']}: {int(payload['rows']):,} rows",
            flush=True,
        )
    elif kind == "complete":
        print(f"Complete in {format_seconds(float(payload['seconds']))}: {payload['file']}", flush=True)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=APP_TITLE)
    parser.add_argument("--input", type=Path, help="每日 D&TMOK KLNG Excel 資料夾")
    parser.add_argument("--output", type=Path, help="輸出 .xlsx")
    parser.add_argument("--legacy-output", type=Path, help="同時輸出的原格式相容版 .xlsx")
    parser.add_argument("--year", type=int)
    parser.add_argument("--month", type=int)
    parser.add_argument("--max-distance", type=float, default=500.0)
    parser.add_argument("--tolerance", type=float, default=10.0, help="百分比，例如 10")
    parser.add_argument("--cluster-size", type=int, default=3)
    parser.add_argument("--top-candidates", type=int, default=50)
    parser.add_argument("--message-types", default="1,2,3,18,19")
    parser.add_argument("--workers", type=int, default=1, help="同時處理的每日檔案數，建議 2–4")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--max-files", type=int, help="測試用：只處理前 N 檔")
    arguments = parser.parse_args(argv)

    if all(value is None for value in (arguments.input, arguments.output, arguments.legacy_output, arguments.year, arguments.month)):
        launch_gui()
        return 0
    if None in (arguments.input, arguments.output, arguments.year, arguments.month):
        parser.error("命令列模式必須同時指定 --input、--output、--year、--month")
    config = AppConfig(
        input_dir=arguments.input,
        output_path=arguments.output,
        year=arguments.year,
        month=arguments.month,
        legacy_output_path=arguments.legacy_output,
        max_distance=arguments.max_distance,
        tolerance=arguments.tolerance / 100.0,
        cluster_size=arguments.cluster_size,
        top_candidates=arguments.top_candidates,
        message_types=parse_message_types(arguments.message_types),
        workers=arguments.workers,
        overwrite=arguments.overwrite,
        max_files=arguments.max_files,
    )
    run_pipeline(config, callback=cli_progress)
    return 0


if __name__ == "__main__":
    multiprocessing.freeze_support()
    raise SystemExit(main())
