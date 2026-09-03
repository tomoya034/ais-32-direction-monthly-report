from __future__ import annotations

import tempfile
import unittest
import shutil
import sys
from io import StringIO
from datetime import date
from pathlib import Path
from unittest import mock

import openpyxl

from ais_monthly_app import (
    AppConfig,
    Candidate,
    _cache_directory,
    _legacy_spool_directory,
    build_processing_job,
    default_output_filename,
    default_output_directory,
    default_worker_count,
    detect_source_catalog,
    derive_legacy_output_path,
    degree_to_direction_index,
    friendly_error_message,
    main,
    month_option_label,
    parse_source_date,
    parse_source_filename,
    process_source_file,
    process_month,
    resolve_source_period,
    resolve_source_month,
    run_pipeline,
    select_cluster,
    select_cluster_from_sorted_rows,
)


class RuleTests(unittest.TestCase):
    def test_filename_date_uses_yyyymmdd_component(self) -> None:
        self.assertEqual(parse_source_date("D&TMOK KLNG_20260401_23.xlsx"), date(2026, 4, 1))
        self.assertIsNone(parse_source_date("KLNG_20260401.xlsx"))

    def test_month_is_derived_from_fixed_filenames_without_manual_input(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK KLNG_20260401_23.xlsx").touch()
            (folder / "D&TMOK KLNG_20260402_23.xlsx").touch()
            (folder / "D&TMOK KLNG_20260501_23.xlsx").touch()
            self.assertEqual(resolve_source_month(folder), (2026, 5, 1))
            self.assertEqual(resolve_source_month(folder, 2026, 4), (2026, 4, 2))
            self.assertEqual(month_option_label(2026, 4, 2), "2026 年 4 月（2 個每日檔）")

    def test_requested_month_must_exist_in_source_filenames(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK KLNG_20260401_23.xlsx").touch()
            with self.assertRaisesRegex(ValueError, "檔名中沒有"):
                resolve_source_month(folder, 2026, 5)

    def test_direction_mapping_matches_existing_script_bins(self) -> None:
        self.assertEqual(degree_to_direction_index(0)[0], 0)
        self.assertEqual(degree_to_direction_index(11.25)[0], 1)
        self.assertEqual(degree_to_direction_index(359.9)[0], 31)
        self.assertEqual(degree_to_direction_index(360)[0], 0)

    def test_cluster_skips_isolated_high_value(self) -> None:
        values = [100.0, 80.0, 79.0, 78.0, 60.0]
        candidates = [Candidate(value, 0.0, index + 2) for index, value in enumerate(values)]
        result = select_cluster("北", candidates, tolerance=0.10, cluster_size=3, over_cap_count=0, over_cap_max=None)
        self.assertEqual(result.selected, 80.0)
        self.assertEqual(result.selected_rank, 2)
        self.assertEqual(result.status, "自動採用")

    def test_coastal_value_is_flagged(self) -> None:
        candidates = [Candidate(value, 247.5, index + 2) for index, value in enumerate((12.0, 11.8, 11.5))]
        result = select_cluster("西南西", candidates, tolerance=0.10, cluster_size=3, over_cap_count=0, over_cap_max=None)
        self.assertEqual(result.status, "待複核")

    def test_full_numeric_search_can_find_cluster_below_saved_top_rows(self) -> None:
        config = AppConfig(
            input_dir=Path("."),
            output_path=Path("result.xlsx"),
            port="KLNG",
            year=2026,
            month=4,
            top_candidates=3,
        )
        rows = [(100.0, 2, 0.0), (80.0, 3, 0.0), (60.0, 4, 0.0), (40.0, 5, 0.0), (20.0, 6, 0.0), (19.0, 7, 0.0), (18.0, 8, 0.0)]
        result = select_cluster_from_sorted_rows("北", rows, config, over_cap_count=0, over_cap_max=None)
        self.assertEqual(result.selected, 20.0)
        self.assertEqual(result.selected_rank, 5)
        self.assertEqual([candidate.rank for candidate in result.candidates], [1, 2, 3, 5, 6, 7])

    def test_output_names_and_common_error_are_friendly(self) -> None:
        modern = Path("KLNG 2026年4月 32方位數值_新版自動分析.xlsx")
        self.assertEqual(derive_legacy_output_path(modern).name, "KLNG 2026年4月 32方位數值_原格式相容版.xlsx")
        self.assertIn("Excel", friendly_error_message(PermissionError("locked")))

    def test_frozen_default_output_stays_beside_portable_exe(self) -> None:
        fake_executable = Path("B:/AIS/AIS_32方位月報工具/AIS_32方位月報工具.exe")
        with (
            mock.patch.object(sys, "frozen", True, create=True),
            mock.patch.object(sys, "executable", str(fake_executable)),
        ):
            self.assertEqual(
                default_output_directory(),
                fake_executable.resolve().parent / "output" / "AIS月報",
            )

    def test_default_workers_are_bounded_for_desktop_use(self) -> None:
        with mock.patch("ais_monthly_app.os.cpu_count", return_value=24):
            self.assertEqual(default_worker_count(), 3)
        with mock.patch("ais_monthly_app.os.cpu_count", return_value=1):
            self.assertEqual(default_worker_count(), 1)
        with mock.patch("ais_monthly_app.os.cpu_count", return_value=None):
            self.assertEqual(default_worker_count(), 1)

    def test_ais_sheet_is_found_after_a_cover_sheet(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            source = folder / "D&TMOK KLNG_20260401_23.xlsx"
            workbook = openpyxl.Workbook()
            cover = workbook.active
            cover.title = "說明"
            cover.append(["這不是 AIS 資料表"])
            sheet = workbook.create_sheet("AIS")
            sheet.append(["msg_type", "LONGITUDE_DESC", "bearing", "distance in nautical miles"])
            sheet.append([1, "East", 0.0, 80.0])
            sheet.append([2, "East", 0.0, 79.0])
            sheet.append([3, "East", 0.0, 78.0])
            workbook.save(source)

            result = process_source_file(
                source,
                date(2026, 4, 1),
                AppConfig(
                    input_dir=folder,
                    output_path=folder / "result.xlsx",
                    port="KLNG",
                    year=2026,
                    month=4,
                ),
            )
            self.assertEqual(result.rows_accepted, 3)
            self.assertEqual(result.directions["北"].selected, 80.0)


class EndToEndTests(unittest.TestCase):
    def test_small_month_build(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            source = folder / "D&TMOK KLNG_20260401_23.xlsx"
            workbook = openpyxl.Workbook(write_only=True)
            sheet = workbook.create_sheet("AIS")
            sheet.append(["msg_type", "LONGITUDE_DESC", "bearing", "distance in nautical miles", "unused"])
            rows = [
                (1, "East", 0.0, 100.0),
                (3, "East", 0.0, 80.0),
                (18, "East", 0.0, 79.0),
                (19, "East", 0.0, 78.0),
                (2, "East", 247.5, 12.0),
                (2, "East", 247.5, 11.8),
                (2, "East", 247.5, 11.5),
                (2, "East", 0.0, 499.0),
                (1, "West", 0.0, 499.0),
                (1, "East", 0.0, 600.0),
            ]
            for row in rows:
                sheet.append([*row, "x"])
            workbook.save(source)
            shutil.copyfile(source, folder / "D&TMOK KLNG_20260402_23.xlsx")

            output = folder / "result.xlsx"
            legacy_output = folder / "result_legacy.xlsx"
            config = AppConfig(
                input_dir=folder,
                output_path=output,
                port="KLNG",
                year=2026,
                month=4,
                legacy_output_path=legacy_output,
                top_candidates=10,
                workers=2,
                overwrite=True,
            )
            run_pipeline(config)
            self.assertTrue(output.exists())
            self.assertTrue(legacy_output.exists())

            events = []
            run_pipeline(config, callback=events.append)
            self.assertTrue(any(event.get("kind") == "cache_hit" for event in events))

            result = openpyxl.load_workbook(output, read_only=False, data_only=False)
            try:
                self.assertIn("總表", result.sheetnames)
                self.assertIn("待複核", result.sheetnames)
                self.assertEqual(result["4月1日"]["H2"].value, 80.0)
                self.assertEqual(result["4月1日"]["AD2"].value, 12.0)
                self.assertIn("'4月1日'!H4", str(result["總表"]["B5"].value))
                self.assertEqual(len(result["總表"]._charts), 2)
                self.assertEqual(result["處理紀錄"]["B1"].value, "KLNG")
                self.assertEqual(result["處理紀錄"]["B2"].value, "2026-04")
                self.assertEqual(result["處理紀錄"]["B3"].value, 2)
            finally:
                result.close()

            legacy = openpyxl.load_workbook(legacy_output, read_only=False, data_only=False)
            try:
                self.assertIn("4月1日", legacy.sheetnames)
                self.assertIn("總表", legacy.sheetnames)
                self.assertIn("工作", legacy.sheetnames)
                self.assertEqual(legacy["4月1日"]["A2"].value, 600.0)
                self.assertEqual(legacy["4月1日"]["C2"].value, "北")
                self.assertEqual(legacy["4月1日"]["H2"].value, 80.0)
                self.assertEqual(legacy["4月1日"]["AD2"].value, 12.0)
                self.assertIn("'4月1日'!H2", str(legacy["總表"]["B2"].value))
                self.assertIsNone(legacy["總表"]["M2"].value)
                self.assertIsNone(legacy["總表"]["M33"].value)
                self.assertEqual(legacy["工作"].sheet_state, "hidden")
            finally:
                legacy.close()


class MultiPortTests(unittest.TestCase):
    @staticmethod
    def _write_sample(path: Path) -> None:
        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet("AIS")
        sheet.append(["msg_type", "LONGITUDE_DESC", "bearing", "distance in nautical miles"])
        sheet.append([1, "East", 0.0, 80.0])
        sheet.append([2, "East", 0.0, 79.0])
        sheet.append([3, "East", 0.0, 78.0])
        workbook.save(path)

    def test_klng_filename_parses_port_and_date(self) -> None:
        parsed = parse_source_filename("D&TMOK KLNG_20260701_23.xlsx")
        self.assertIsNotNone(parsed)
        self.assertEqual((parsed.port, parsed.day), ("KLNG", date(2026, 7, 1)))

    def test_hwln_filename_parses_case_insensitively_and_normalizes_port(self) -> None:
        parsed = parse_source_filename("d&tmok hwln_20260101_23.XLSX")
        self.assertIsNotNone(parsed)
        self.assertEqual((parsed.port, parsed.day), ("HWLN", date(2026, 1, 1)))

    def test_port_pattern_rejects_unsafe_or_overly_loose_names(self) -> None:
        invalid = (
            "D&TMOK A_20260101_23.xlsx",
            "D&TMOK HW-LN_20260101_23.xlsx",
            "D&TMOK HW_LN_20260101_23.xlsx",
            "D&TMOK ../HWLN_20260101_23.xlsx",
            "D&TMOK HWLN_20260101_23 bad.xlsx",
        )
        for filename in invalid:
            with self.subTest(filename=filename):
                self.assertIsNone(parse_source_filename(filename))

    def test_same_port_multiple_days_form_one_period(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            for day in (1, 2, 3):
                (folder / f"D&TMOK HWLN_202601{day:02d}_23.xlsx").touch()
            self.assertEqual(detect_source_catalog(folder), {"HWLN": [(2026, 1, 3)]})

    def test_same_port_multiple_months_stay_separate(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK HWLN_20260101_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260201_23.xlsx").touch()
            self.assertEqual(detect_source_catalog(folder)["HWLN"], [(2026, 1, 1), (2026, 2, 1)])

    def test_multiple_ports_are_separate_and_expose_port_specific_months(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK KLNG_20260601_23.xlsx").touch()
            (folder / "D&TMOK KLNG_20260701_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260101_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260201_23.xlsx").touch()
            catalog = detect_source_catalog(folder)
            self.assertEqual(list(catalog), ["HWLN", "KLNG"])
            self.assertEqual(catalog["KLNG"], [(2026, 6, 1), (2026, 7, 1)])
            self.assertEqual(catalog["HWLN"], [(2026, 1, 1), (2026, 2, 1)])

    def test_unique_port_and_latest_month_are_automatic(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK HWLN_20260101_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260201_23.xlsx").touch()
            self.assertEqual(resolve_source_period(folder), ("HWLN", 2026, 2, 1))

    def test_cli_multi_port_without_port_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK KLNG_20260101_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260101_23.xlsx").touch()
            stderr = StringIO()
            with mock.patch("sys.stderr", stderr), self.assertRaises(SystemExit) as raised:
                main(["--input", str(folder), "--output", str(folder / "result.xlsx")])
            self.assertEqual(raised.exception.code, 2)
            self.assertIn("--port", stderr.getvalue())

    def test_cli_port_selects_only_that_port(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            (folder / "D&TMOK KLNG_20260101_23.xlsx").touch()
            (folder / "D&TMOK HWLN_20260101_23.xlsx").touch()
            with mock.patch("ais_monthly_app.run_pipeline") as mocked_run:
                exit_code = main(
                    [
                        "--input", str(folder),
                        "--output", str(folder / "result.xlsx"),
                        "--port", "hwln",
                    ]
                )
            self.assertEqual(exit_code, 0)
            config = mocked_run.call_args.args[0]
            self.assertEqual((config.port, config.year, config.month), ("HWLN", 2026, 1))

    def test_output_names_are_port_aware(self) -> None:
        modern = Path(default_output_filename("hwln", 2026, 1))
        self.assertEqual(modern.name, "HWLN_2026年01月_32方位數值_新版自動分析.xlsx")
        self.assertEqual(
            derive_legacy_output_path(modern).name,
            "HWLN_2026年01月_32方位數值_原格式相容版.xlsx",
        )

    def test_cache_and_legacy_spool_are_scoped_by_port_and_period(self) -> None:
        common = dict(input_dir=Path("."), output_path=Path("result.xlsx"), year=2026, month=1)
        klng = AppConfig(port="KLNG", **common)
        hwln = AppConfig(port="HWLN", **common)
        self.assertNotEqual(_cache_directory(klng), _cache_directory(hwln))
        self.assertNotEqual(_legacy_spool_directory(klng), _legacy_spool_directory(hwln))

    def test_hwln_pipeline_uses_same_research_results_without_mixing_klng(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as temporary:
            folder = Path(temporary)
            klng_source = folder / "D&TMOK KLNG_20260101_23.xlsx"
            hwln_source = folder / "D&TMOK HWLN_20260101_23.xlsx"
            self._write_sample(klng_source)
            shutil.copyfile(klng_source, hwln_source)
            klng_config = AppConfig(
                input_dir=folder,
                output_path=folder / "klng.xlsx",
                port="KLNG",
                year=2026,
                month=1,
            )
            hwln_output = folder / default_output_filename("HWLN", 2026, 1)
            hwln_config = AppConfig(
                input_dir=folder,
                output_path=hwln_output,
                port="HWLN",
                year=2026,
                month=1,
                overwrite=True,
            )
            job = build_processing_job(hwln_config)
            self.assertEqual((job.port, job.year, job.month), ("HWLN", 2026, 1))
            self.assertEqual([path.name for _day, path in job.files], [hwln_source.name])
            klng_day = process_source_file(klng_source, date(2026, 1, 1), klng_config)
            hwln_day = process_source_file(hwln_source, date(2026, 1, 1), hwln_config)
            self.assertEqual(
                [klng_day.directions[name].selected for name in klng_day.directions],
                [hwln_day.directions[name].selected for name in hwln_day.directions],
            )

            days, _warnings = process_month(hwln_config)
            actual_sources = [item.source_file.name for item in days if item.source_file is not None]
            self.assertEqual(actual_sources, [hwln_source.name])
            run_pipeline(hwln_config)
            workbook = openpyxl.load_workbook(hwln_output, read_only=True)
            try:
                self.assertEqual(workbook.properties.title, "HWLN 2026年01月 AIS 32方位數值")
                self.assertEqual(workbook["操作說明"]["B4"].value, "HWLN")
                self.assertEqual(workbook["處理紀錄"]["B1"].value, "HWLN")
                self.assertEqual(workbook["處理紀錄"]["B3"].value, 1)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
